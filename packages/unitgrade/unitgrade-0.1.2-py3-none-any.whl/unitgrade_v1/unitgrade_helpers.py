import numpy as np
from tabulate import tabulate
from datetime import datetime
import pyfiglet
from unitgrade_v1 import Hidden, myround, msum, ActiveProgress
# import unitgrade_v1

# from unitgrade_v1.unitgrade_v1 import Hidden
# import unitgrade_v1 as ug
# import unitgrade_v1.unitgrade_v1 as ug
import inspect
import os
import argparse
import sys
import time

#from threading import Thread  # This import presents a problem for the minify-code compression tool.

parser = argparse.ArgumentParser(description='Evaluate your report.', epilog="""Example: 
To run all tests in a report: 

> python assignment1_dp.py

To run only question 2 or question 2.1

> python assignment1_dp.py -q 2
> python assignment1_dp.py -q 2.1

Note this scripts does not grade your report. To grade your report, use:

> python report1_grade.py

Finally, note that if your report is part of a module (package), and the report script requires part of that package, the -m option for python may be useful.
For instance, if the report file is in Documents/course_package/report3_complete.py, and `course_package` is a python package, then change directory to 'Documents/` and run:

> python -m course_package.report1

see https://docs.python.org/3.9/using/cmdline.html
""", formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('-q', nargs='?', type=str, default=None, help='Only evaluate this question (e.g.: -q 2)')
parser.add_argument('--showexpected',  action="store_true",  help='Show the expected/desired result')
parser.add_argument('--showcomputed',  action="store_true",  help='Show the answer your code computes')
parser.add_argument('--unmute',  action="store_true",  help='Show result of print(...) commands in code')
parser.add_argument('--passall',  action="store_true",  help='Automatically pass all tests. Useful when debugging.')

def evaluate_report_student(report, question=None, qitem=None, unmute=None, passall=None, ignore_missing_file=False, show_tol_err=False):
    args = parser.parse_args()
    if question is None and args.q is not None:
        question = args.q
        if "." in question:
            question, qitem = [int(v) for v in question.split(".")]
        else:
            question = int(question)

    if not os.path.isfile(report.computed_answers_file) and not ignore_missing_file:
        raise Exception("> Error: The pre-computed answer file", os.path.abspath(report.computed_answers_file), "does not exist. Check your package installation")

    if unmute is None:
        unmute = args.unmute
    if passall is None:
        passall = args.passall

    results, table_data = evaluate_report(report, question=question, show_progress_bar=not unmute and not args.noprogress, qitem=qitem, verbose=False, passall=passall, show_expected=args.showexpected, show_computed=args.showcomputed,unmute=unmute,
                                          show_tol_err=show_tol_err)

    try:  # For registering stats.
        import unitgrade_private_v1
        import irlc.lectures
        import xlwings
        from openpyxl import Workbook
        import pandas as pd
        from collections import defaultdict
        dd = defaultdict(lambda: [])
        error_computed = []
        for k1, (q, _) in enumerate(report.questions):
            for k2, item in enumerate(q.items):
                dd['question_index'].append(k1)
                dd['item_index'].append(k2)
                dd['question'].append(q.name)
                dd['item'].append(item.name)
                dd['tol'].append(0 if not hasattr(item, 'tol') else item.tol)
                error_computed.append(0 if not hasattr(item, 'error_computed') else item.error_computed)

        qstats = report.wdir + "/" + report.name + ".xlsx"

        if os.path.isfile(qstats):
            d_read = pd.read_excel(qstats).to_dict()
        else:
            d_read = dict()

        for k in range(1000):
            key = 'run_'+str(k)
            if key in d_read:
                dd[key] = list(d_read['run_0'].values())
            else:
                dd[key] = error_computed
                break

        workbook = Workbook()
        worksheet = workbook.active
        for col, key in enumerate(dd.keys()):
            worksheet.cell(row=1, column=col+1).value = key
            for row, item in enumerate(dd[key]):
                worksheet.cell(row=row+2, column=col+1).value = item

        workbook.save(qstats)
        workbook.close()

    except ModuleNotFoundError as e:
        s = 234
        pass

    if question is None:
        print("Provisional evaluation")
        tabulate(table_data)
        table = table_data
        print(tabulate(table))
        print(" ")

    fr = inspect.getouterframes(inspect.currentframe())[1].filename
    gfile = os.path.basename(fr)[:-3] + "_grade.py"
    if os.path.exists(gfile):
        print("Note your results have not yet been registered. \nTo register your results, please run the file:")
        print(">>>", gfile)
        print("In the same manner as you ran this file.")
    return results


def upack(q):
    # h = zip([(i['w'], i['possible'], i['obtained']) for i in q.values()])
    h =[(i['w'], i['possible'], i['obtained']) for i in q.values()]
    h = np.asarray(h)
    return h[:,0], h[:,1], h[:,2],



def evaluate_report(report, question=None, qitem=None, passall=False, verbose=False,  show_expected=False, show_computed=False,unmute=False, show_help_flag=True, silent=False,
                    show_progress_bar=True,
                    show_tol_err=False):
    from src.snipper.version import __version__
    now = datetime.now()
    ascii_banner = pyfiglet.figlet_format("UnitGrade", font="doom")
    b = "\n".join( [l for l in ascii_banner.splitlines() if len(l.strip()) > 0] )
    print(b + " v" + __version__)
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    print("Started: " + dt_string)
    s = report.title
    if report.version is not None:
        s += " version " + report.version
    print("Evaluating " + s, "(use --help for options)" if show_help_flag else "")
    print(f"Loaded answers from: ", report.computed_answers_file, "\n")
    table_data = []
    nL = 80
    t_start = time.time()
    score = {}
    for n, (q, w) in enumerate(report.questions):
        q_hidden = issubclass(q.__class__, Hidden)
        if question is not None and n+1 != question:
            continue
        q_title_print = "Question %i: %s"%(n+1, q.title)
        print(q_title_print, end="")
        q.possible = 0
        q.obtained = 0
        q_ = {} # Gather score in this class.

        q_with_outstanding_init = [item.question for item in q.items if not item.question.has_called_init_]

        for j, item in enumerate(q.items):
            if qitem is not None and question is not None and j+1 != qitem:
                continue

            if q_with_outstanding_init is not None: # check for None bc. this must be called to set titles.
                # if not item.question.has_called_init_:
                start = time.time()

                cc = None
                if show_progress_bar:
                    total_estimated_time = q.estimated_time # Use this. The time is estimated for the q itself.  # sum( [q2.estimated_time for q2 in q_with_outstanding_init] )
                    cc = ActiveProgress(t=total_estimated_time, title=q_title_print)
                with eval('Capturing')(unmute=unmute):  # Clunky import syntax is required bc. of minify issue.
                    try:
                        for q2 in q_with_outstanding_init:
                            q2.init()
                            q2.has_called_init_ = True

                        # item.question.init()  # Initialize the question. Useful for sharing resources.
                    except Exception as e:
                        if not passall:
                            if not silent:
                                print(" ")
                                print("="*30)
                                print(f"When initializing question {q.title} the initialization code threw an error")
                                print(e)
                                print("The remaining parts of this question will likely fail.")
                                print("="*30)

                if show_progress_bar:
                    cc.terminate()
                    sys.stdout.flush()
                    print(q_title_print, end="")

                # item.question.has_called_init_ = True
                q_time =np.round(  time.time()-start, 2)

                print(" "* max(0,nL - len(q_title_print) ) + (" (" + str(q_time) + " seconds)" if q_time >= 0.1 else "") ) # if q.name in report.payloads else "")
                print("=" * nL)
                q_with_outstanding_init = None

            # item.question = q # Set the parent question instance for later reference.
            item_title_print = ss = "*** q%i.%i) %s"%(n+1, j+1, item.title)

            if show_progress_bar:
                cc = ActiveProgress(t=item.estimated_time, title=item_title_print)
            else:
                print(item_title_print + ( '.'*max(0, nL-4-len(ss)) ), end="")
            hidden = issubclass(item.__class__, Hidden)
            # if not hidden:
            #     print(ss, end="")
            # sys.stdout.flush()
            start = time.time()

            (current, possible) = item.get_points(show_expected=show_expected, show_computed=show_computed,unmute=unmute, passall=passall, silent=silent)
            q_[j] = {'w': item.weight, 'possible': possible, 'obtained': current, 'hidden': hidden, 'computed': str(item._computed_answer), 'title': item.title}
            tsecs = np.round(time.time()-start, 2)
            if show_progress_bar:
                cc.terminate()
                sys.stdout.flush()
                print(item_title_print + ('.' * max(0, nL - 4 - len(ss))), end="")

            if not hidden:
                ss = "PASS" if current == possible else "*** FAILED"
                if tsecs >= 0.1:
                    ss += " ("+ str(tsecs) + " seconds)"
                print(ss)

        ws, possible, obtained = upack(q_)
        possible = int(ws @ possible)
        obtained = int(ws @ obtained)
        obtained = int(myround(int((w * obtained) / possible ))) if possible > 0 else 0
        score[n] = {'w': w, 'possible': w, 'obtained': obtained, 'items': q_, 'hidden': q_hidden, 'title': q.title}

        q.obtained = obtained
        q.possible = possible

        s1 = f"*** Question q{n+1}"
        s2 = f" {q.obtained}/{w}"
        print(s1 + ("."* (nL-len(s1)-len(s2) )) + s2 )
        print(" ")
        table_data.append([f"Question q{n+1}", f"{q.obtained}/{w}"])

    ws, possible, obtained = upack(score)
    possible = int( msum(possible) )
    obtained = int( msum(obtained) ) # Cast to python int
    report.possible = possible
    report.obtained = obtained
    now = datetime.now()
    dt_string = now.strftime("%H:%M:%S")

    dt = int(time.time()-t_start)
    minutes = dt//60
    seconds = dt - minutes*60
    plrl = lambda i, s: str(i) + " " + s + ("s" if i != 1 else "")

    print(f"Completed: "+ dt_string + " (" + plrl(minutes, "minute") + ", "+ plrl(seconds, "second") +")")

    table_data.append(["Total", ""+str(report.obtained)+"/"+str(report.possible) ])
    results = {'total': (obtained, possible), 'details': score}
    return results, table_data
