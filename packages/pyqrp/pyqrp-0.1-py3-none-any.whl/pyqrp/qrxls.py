"""
qrxls.py - Part of PyQRP

:author: George <drpresq@gmail.com>
:description: PyQRP XLS Module - Handles the parsing of XLS and XLSX files
:license: GPL3
:donation:
    BTC - 15wRP3NGm2zQwsC36gYAMf8ZaBNuDP6BiR
    LTC - LQANeFg6qhEUCftCGpXTdgCKnPkBMR5Ems
"""

import openpyxl
import xlrd
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional, Union


class QRXls:
    """
    QRXlx - xlrd Based Pilot XLS report parser
    """
    __xls_path: Path
    __xls_file: xlrd.Book
    __xls_sheets: list
    __target_sheet: dict

    def __init__(self, file_path: Optional[Union[str, Path]] = None) -> None:
        self.__xls_path = \
            Path(file_path) if isinstance(file_path, str) and Path(file_path).exists() else \
            file_path if isinstance(file_path, Path) and file_path.exists() else None

        if self.__xls_path:
            self._parse_file()

    def _is_header_row(self, row: int, col: int, sheet: xlrd.sheet) -> bool:
        if col < sheet.ncols and sheet.cell(row, col):
            return self._is_header_row(row, col + 1, sheet)
        elif col < sheet.ncols and not sheet.cell(row, col):
            return False
        else:
            return True

    def _parse_file(self, target_sheet_index: Optional[int] = None) -> None:
        self.__xls_sheets = [] if not isinstance(self.__xls_sheets, list) else self.__xls_sheets
        self.__xls_file = xlrd.open_workbook(self.__xls_path)

        # For each sheet in the workbook
        for sheet in self.__xls_file.sheets():
            if target_sheet_index and (sheet.number != target_sheet_index and sheet.name != target_sheet_index):
                continue
            sheet_headers: list = []
            out_sheet: list = []
            row: int = 0

            # Find the data headers
            while row < sheet.nrows:
                if self._is_header_row(row=row, col=0, sheet=sheet):
                    break
                row += 1
            else:
                raise Exception("I couldn't find any data headers.")
            # Record data headers
            for col in sheet.ncols:
                sheet_headers.append(sheet.cell(row, col))

            # Process the data row by row
            for row in range(0, sheet.nrows):
                if sheet.cell(row, 0) and sheet.cell(row, 0).value != sheet_headers[0]:
                    row_data: dict = {}
                    for header in sheet_headers:
                        row_data.update({header: sheet.cell(row, sheet_headers.index(header))})
                    out_sheet.append(row_data)

            self.__xls_sheets.append({"name": sheet.name,
                                      "index": sheet.number,
                                      "data": out_sheet})

    def __iter__(self):
        self.__n = 0
        return self

    def __next__(self):
        if self.__n <= len(self.__xls_sheets):
            ret = self.__xls_sheets[self.__n]
            self.__n += 1
            return ret
        else:
            raise StopIteration

    @property
    def sheets(self) -> list:
        return self.__xls_file.sheets()

    @property
    def target_sheet(self) -> dict:
        return self.target_sheet

    @target_sheet.setter
    def target_sheet(self, target_sheet_index: Optional[Union[str, int]] = None) -> None:
        self.target_sheet = \
            target_sheet_index if target_sheet_index else \
            self.target_sheet if self.target_sheet else {}

    @property
    def book_data(self) -> list:
        if not self.__xls_sheets:
            self._parse_file(self.target_sheet if self.target_sheet else None)
        return self.__xls_sheets


class QRXlsx:
    """
    QRXlxs - Openpyxl Based Pilot XLSX report parser
    """
    __xlsx_path: Path
    __xlsx_file: openpyxl.Workbook
    __xlsx_sheets: list
    __target_sheet: dict

    def __init__(self, file_path: Union[str, Path]) -> None:
        self.__xlsx_path = \
            Path(file_path) if isinstance(file_path, str) and Path(file_path).exists() else \
            file_path if isinstance(file_path, Path) and file_path.exists() else None

        if self.__xlsx_path:
            self._parse_file()
        else:
            raise FileNotFoundError

    def _is_header_row(self, row: int, col: int, sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        if col <= sheet.max_column and sheet.cell(row, col).value != '':
            return self._is_header_row(row, col + 1, sheet)
        elif col <= sheet.max_column and not sheet.cell(row, col).value != '':
            return False
        else:
            return True

    def _parse_file(self, target_sheet_name: Optional[str] = None) -> None:
        try:
            self.__xlsx_sheets = [] if not self.__xlsx_sheets else self.__xlsx_sheets
        except AttributeError:
            self.__xlsx_sheets = []

        self.__xlsx_file = openpyxl.open(self.__xlsx_path, data_only=True)

        # For each sheet in the workbook
        for sheet in self.__xlsx_file.worksheets:
            if target_sheet_name and sheet._WorkbookChild__title != target_sheet_name:
                continue
            sheet_headers: list = []
            out_sheet: list = []
            row: int = 1

            # Find the data headers
            while row <= sheet.max_row:
                if self._is_header_row(row=row, col=1, sheet=sheet):
                    break
                row += 1
            else:
                raise Exception("I couldn't find any data headers.")
            # Record data headers
            for col in range(1, sheet.max_column+1):
                sheet_headers.append(sheet.cell(row, col).value)

            # Process the data row by row
            for row in range(1, sheet.max_row+1):
                if sheet.cell(row, 1).value and sheet.cell(row, 1).value != sheet_headers[0]:
                    row_data: dict = {}
                    for header in sheet_headers:
                        row_data.update({header: sheet.cell(row, sheet_headers.index(header)+1).value})
                    out_sheet.append(row_data)

            self.__xlsx_sheets.append({"name": sheet._WorkbookChild__title,
                                       "index": self.__xlsx_file.sheetnames.index(sheet._WorkbookChild__title)+1,
                                       "data": out_sheet})

    def __iter__(self):
        self.__n = 0
        return self

    def __next__(self):
        if self.__n <= len(self.__xlsx_sheets):
            ret = self.__xlsx_sheets[self.__n]
            self.__n += 1
            return ret
        else:
            raise StopIteration

    @property
    def sheets(self) -> list:
        return self.__xlsx_file.sheetnames

    @property
    def target_sheet(self) -> dict:
        return self.target_sheet

    @target_sheet.setter
    def target_sheet(self, target_sheet_name: Optional[str] = None) -> None:
        self.target_sheet = \
            target_sheet_name if target_sheet_name else \
            self.target_sheet if self.target_sheet else {}

    @property
    def book_data(self) -> list:
        if not self.__xlsx_sheets:
            self._parse_file(self.target_sheet if self.target_sheet else None)
        return self.__xlsx_sheets
