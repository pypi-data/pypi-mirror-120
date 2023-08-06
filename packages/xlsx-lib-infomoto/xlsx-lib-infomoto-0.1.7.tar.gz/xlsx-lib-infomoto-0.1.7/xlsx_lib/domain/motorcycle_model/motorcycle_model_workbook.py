from typing import Optional, List

from openpyxl import \
    Workbook, \
    load_workbook

from xlsx_lib.domain.frame.frame_element import FrameElement
from xlsx_lib.domain.frame.frame_sheet import FrameSheet
from xlsx_lib.domain.motorcycle_model.motorcycle_model import MotorcycleModel
from xlsx_lib.domain.motorcycle_model.sheetnames import Sheetnames

from xlsx_lib.domain.engine.engine_section import EngineSection
from xlsx_lib.domain.engine.engine_sheet import EngineSheet

from xlsx_lib.domain.electronic.electronic_element import ElectronicElement
from xlsx_lib.domain.electronic.electronic_sheet import ElectronicSheet

from xlsx_lib.domain.generic_replacements.generic_replacement_sheet import GenericReplacementsSheet
from xlsx_lib.domain.generic_replacements.replacement import Replacement

from xlsx_lib.domain.tightening_torques.tightening_torques_sheet import TighteningTorquesSheet
from xlsx_lib.domain.tightening_torques.component import Component

sheetnames_relationships = {
    "MOT": Sheetnames.ENGINE,
    "REC. GENERICOS": Sheetnames.GENERIC_REPLACEMENTS,
    "ELEC": Sheetnames.ELECTRONIC,
    "PARES APRIETE": Sheetnames.TIGHTENING_TORQUES,
    "CHAS": Sheetnames.FRAME,
}


class MotorcycleModelWorkbook:
    motorcycle_model: MotorcycleModel

    def __init__(self, filename: any, directory: Optional[str] = None):
        workbook: Workbook = load_workbook(filename=filename)

        generic_replacements: Optional[List[Replacement]] = None
        electronic_elements: Optional[List[ElectronicElement]] = None
        components_screws_tightening_torques: Optional[List[Component]] = None
        engine_sections: Optional[List[EngineSection]] = None
        frame_elements: Optional[List[FrameElement]] = None

        for sheetname in [key for key in workbook.sheetnames if key in sheetnames_relationships]:
            if sheetnames_relationships[sheetname] == Sheetnames.ENGINE:
                engine_sections = \
                    EngineSheet(worksheet=workbook[sheetname]).get_electronic_elements()
            elif sheetnames_relationships[sheetname] == Sheetnames.GENERIC_REPLACEMENTS:
                generic_replacements = \
                    GenericReplacementsSheet(worksheet=workbook[sheetname]).get_generic_replacements()
            elif sheetnames_relationships[sheetname] == Sheetnames.ELECTRONIC:
                electronic_elements = \
                    ElectronicSheet(worksheet=workbook[sheetname]).get_electronic_elements()
            elif sheetnames_relationships[sheetname] == Sheetnames.TIGHTENING_TORQUES:
                components_screws_tightening_torques = \
                    TighteningTorquesSheet(worksheet=workbook[sheetname]).get_components_screws_tightening_torques()
            elif sheetnames_relationships[sheetname] == Sheetnames.FRAME:
                frame_elements = FrameSheet(worksheet=workbook[sheetname]).frame_elements

        self.motorcycle_model = MotorcycleModel(
            model_name=filename[filename.rfind("/") + 1:filename.rfind(".")].rstrip().replace("FICHA ", ""),
            generic_replacements=generic_replacements,
            components_screws_tightening_torques=components_screws_tightening_torques,
            electronic_elements=electronic_elements,
            engine_sections=engine_sections,
            frame_elements=frame_elements,
            directory_name=directory,
        )
